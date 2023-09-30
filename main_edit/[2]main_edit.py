# IMPORTS...................................
import logging
import traceback

from bs4 import BeautifulSoup
import requests
import selenium.common.exceptions
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from math import floor
from datetime import datetime, date
import os

# "Global" objects...........................

# global save lists:
all_names_save_list = []
all_prices_save_list = []
all_prices_sub_save_list = []

# list with premade file names (expandable)
possible_spreadsheets_list = ["C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[3]any_run_saves\\webscraper_test.xlsx", "C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[3]any_run_saves\\onlycases.xlsx"]

# dictionary's and lists for global item accessibility
GLOBAL_PARAM_DICT = {"exceptions": False, "daily_save": False, "log_file_index": 0}
SEARCH_PARAM_DICT = {}
PARAMS_LIST = ["daily_save", "only_cases", "limit", "sort", "order", "blacklist", "search", "iterations", "calc_dmarket_steam", "daily_save"]
ACTUAL_PARAM_LIST = []

# global values
WAIT_TIME = 10
LOG_FILE_INDEX = 0

# default translation table
TRANSLATION_TABLE = dict.fromkeys(map(ord, '\n\t'), None)

# datetime at the start of the programm
DATETIME = datetime.today()
DATE = date.today()


# FUNCTIONS...................................

# / main function:
# params(log file index)
#
# \#
def main(lfi):

    # creating error list index
    i = 0

    while not i == 20:  # 20 because there can be only 20 log files per day
                        # function searches which log file names already exist and takes one which doesn't

        lfi += 1

        if os.path.exists(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[{str(lfi)}].txt") is True:

            pass

        elif os.path.exists(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[1]result_files\\results_for_{str(DATE)}[{str(lfi)}].txt") is False:

            break

        elif i == 20:

            print(f"Aborting log file. There can't be more than 20.")

        i += 1

    # saves the log file index of the current run
    GLOBAL_PARAM_DICT["log_file_index"] = int(lfi)

    # adds log text 1 (start time of log)
    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[{str(lfi)}].txt", "x", encoding="utf-8") as lf:

        lf.write(f"Log file {lfi} for normal run\n"
                 f"Starting log at {datetime.today()}\n")

    # parameters:
    only_cases = input("Search only for cases? (y/n) ").lower()  # decides if to only search for cases or not
    limit = input("Price limit? (integer) ").lower()  # decides the price limit of the search
    sort = input("Sort by ...? (buff, cheapest, marketcap, tradevolume) ").lower()  # decides how the search items are sorted
    order = input("Sort ascending or descending? (a/d) ").lower()  # decides if search items are ASC or DESC
    blacklist = input("Blacklist items: (comma seperated, no space) ").lower()  # blacklists characters
    search = input("Search for (a) specific item/s: ").lower()  # takes specific search arguments

    if only_cases == "" or only_cases == "n":

        iterations = input("How many pages to scrape? ").lower()  # decides how many pages will be scraped

    else:

        iterations = ""

    calc_dmarket_steam = input("Calculate dmarket/steam arbitrage? It takes long. (y/n) ")  # decides if the dmarket/steam arbitrage should be calculated for each item
    daily_save = input("Is this a daily save for historical data? (y/n) ")  # decides if run is treated as daily save

    # default params:
    if limit == "":
        limit = 1000000

    if sort == "":
        sort = "marketcap"

    if order == "":
        order = ":DESC"

    elif order == "a":
        order = ":ASC"

    elif order == "d":
        order = ":DESC"

    if only_cases == "":
        only_cases = "n"

    if iterations == "":

        if only_cases == "y":

            iterations = 2

        else:

            iterations = 10

    if calc_dmarket_steam == "":
        calc_dmarket_steam = False

    if daily_save == "":

        daily_save = "n"

    elif daily_save == "y":

        GLOBAL_PARAM_DICT["daily_save"] = True

    # assign sheet index:
    if only_cases == "y":

        sheet_index = 1

    else:
        sheet_index = 0

    # print all info
    print(f"\nonly cases: {only_cases}\n"
          f"limit: {limit}\n"
          f"sorting: {sort}\n"
          f"ordering: {order}\n"
          f"iterations(pages): {iterations}\n"
          f"...\n"
          f"Everything Ok\n"
          f"Expected run-time: {round(float(((int(iterations) * 10) + 40) / 60), 1)} min. (+- 30 sec if error-free)\n"
          f"Expected amount of items: {int(int(iterations)*30*4)}\n"
          )

    print("...")
    print(f"Getting html contents...")

    ACTUAL_PARAM_LIST.append(daily_save)
    ACTUAL_PARAM_LIST.append(only_cases)
    ACTUAL_PARAM_LIST.append(limit)
    ACTUAL_PARAM_LIST.append(sort)
    ACTUAL_PARAM_LIST.append(order)
    ACTUAL_PARAM_LIST.append(blacklist)
    ACTUAL_PARAM_LIST.append(search)
    ACTUAL_PARAM_LIST.append(iterations)
    ACTUAL_PARAM_LIST.append(calc_dmarket_steam)

    for i in range(len(PARAMS_LIST)):

        try:

            SEARCH_PARAM_DICT[i - 1] = str(PARAMS_LIST[i])

        except:

            SEARCH_PARAM_DICT.update({i - 1: str(PARAMS_LIST[i])})

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[{str(lfi)}].txt", "a", encoding="utf-8") as lf:

        lf.write(f"Params assigned at {datetime.today()}: \n\n")

        for i in range(0, len(PARAMS_LIST) - 1):
            lf.write(f"{str(PARAMS_LIST[i])} --- {str(ACTUAL_PARAM_LIST[i])}\n")

        lf.write("\n")

    try:
        # functions:

        # get the html contents

        get_html(itr=iterations, oc=only_cases, first=True, limit=limit, sort=sort, order=order, bl=blacklist,
                 search=search, ds=daily_save)

        # make a spreadsheet with all the dat

        print_results_yn = input("Print result text file? (y/n) ").lower()

        if print_results_yn == "":

            print_results_yn = False

        elif print_results_yn == "y":

            print_results_yn = True

        elif print_results_yn == "n":

            print_results_yn = False

        if print_results_yn is True:

            print_results(
                si=make_spreadsheet_normal(an=all_names_save_list, aps=all_prices_sub_save_list, cds=calc_dmarket_steam,
                                           si=sheet_index))

            with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                    GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

                lf.write(f"completed programm without errors at {datetime.today()}\n")

        else:

            make_spreadsheet_normal(an=all_names_save_list, aps=all_prices_sub_save_list, cds=calc_dmarket_steam,
                                    si=sheet_index)

            with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                    GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

                lf.write(f"completed programm without errors at {datetime.today()}\n")
    except Exception as e:

        print(f"Error when running programm")
        with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a") as lf:

            lf.write(f"did not complete programm regularly     at {datetime.today()}\n")

        print(e)
        # print(traceback.print_exc())


# get html contents of page and start extraction function
def get_html(itr, oc, first, limit, sort, order, bl, search, ds):
    accesed_website = ""
    # driver (currently Edge):
    driver = webdriver.Edge()

    # check if only cases, then go to site and wait for page load
    if oc == "y":

        print(
            f"accessing Website at https://pricempire.com/trending?to={limit}&sort={sort}{order}&blacklist=hardened&search=case...")

        driver.get(
            f"https://pricempire.com/trending?to={limit}&sort={sort}{order}&blacklist=hardened&search=case")

        try:

            driver.find_element(By.CSS_SELECTOR, ".fc-cta-consent").click()

        except:

            print("Button not dound")

        accesed_website = f"https://pricempire.com/trending?to={limit}&sort={sort}{order}&blacklist=hardened&search=case"
    else:

        print(
            f"accessing Website at https://pricempire.com/trending?to={limit}&sort={sort}{order}&blacklist={bl}&search={search}...")

        driver.get(f"https://pricempire.com/trending?to={limit}&sort={sort}{order}&blacklist={bl}&search={search}")

        try:

            driver.find_element(By.CSS_SELECTOR, ".fc-cta-consent").click()

        except:

            print("Button not dound")

        accesed_website = f"https://pricempire.com/trending?to={limit}&sort={sort}{order}&blacklist={bl}&search={search}"

    # write to log file
    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a",
              encoding="utf-8") as lf:

        lf.write(f"accesed website ({accesed_website}) at {datetime.today()}     at {datetime.today()}\n")

    # page load time:
    print("...")
    time.sleep(WAIT_TIME)

    # get html contents of all pages by clicking next page button and start extraction function
    for i in range(int(itr)):

        print(f"Getting Page ({i+1}) contents...")

        with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a",
                  encoding="utf-8") as lf:

            lf.write(f"Getting Page ({i + 1 }) contents...       at {datetime.today()}... ")

        if first is True:

            # get contents
            contents = BeautifulSoup(driver.page_source, "html.parser")

            print(f"extracting data...")

            # extract the data into a list:
            extract_data(contents)

            # alter so next time it goes through "else":
            first = False

        else:

            # click next page button and wait for page load:

            try:

                driver.find_element(By.CSS_SELECTOR, ".fas.fa-chevron-right").click()
                time.sleep(WAIT_TIME)

                # get contents:
                contents = BeautifulSoup(driver.page_source, "html.parser")

                print(f"extracting data...")

                # extract the data into a list
                if extract_data(contents) is False:

                    # click next page button and wait for page load longer this time:

                    driver.refresh()
                    time.sleep(WAIT_TIME + 10)

                    # get contents:
                    contents_reload = BeautifulSoup(driver.page_source, "html.parser")

                    print(f"extracting data...")

                    if extract_data(contents_reload) is False:

                        with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                                GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:
                            lf.write(f"UNSUCCESFULL     at {datetime.today()}\n")
                            lf.write(f"attempting to skip to next page...     at {datetime.today()}... ")

                        print(f"attempting to skip to next page... ")

                        # check if only cases, then go to site and wait for page load
                        if oc == "y":

                            print(
                                f"accessing Website at https://pricempire.com/trending?page={itr}&to={limit}&sort={sort}{order}&blacklist=hardened&search=case...")

                            driver.get(
                                f"https://pricempire.com/trending?page={itr}&to={limit}&sort={sort}{order}&blacklist=hardened&search=case")

                        else:

                            print(
                                f"accessing Website at https://pricempire.com/trending?page={itr}&to={limit}&sort={sort}{order}&blacklist={bl}&search={search}...")

                            driver.get(
                                f"https://pricempire.com/trending?page={itr}&to={limit}&sort={sort}{order}&blacklist={bl}&search={search}")

                        contents_reload_2 = BeautifulSoup(driver.page_source, "html.parser")

                        if extract_data(contents_reload) is False:
                            with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                                    GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:
                                lf.write(f"UNSUCCESSFUL     at {datetime.today()}\n")
                                lf.write(
                                    f"aborting and saving current items (amount: {str(len(all_names_save_list))})     at {datetime.today()}\n")

                            print(
                                f"Something broke, saving current items (amount: {str(len(all_names_save_list))}) and aborting")

                            break
                else:
                    pass

            except selenium.common.exceptions.NoSuchElementException:

                with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                        GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:
                    lf.write(f"UNSUCCESSFUL     at {datetime.today()}\n")
                    lf.write(f"Already at last page, going to make spreadsheet now     at {datetime.today()}\n")

                GLOBAL_PARAM_DICT["exceptions"] = True

                print("Exception: {} Already at last page, going to make spreadsheet now".format(
                    GLOBAL_PARAM_DICT["exceptions"]))

                break


# extract the data by stripping all unneeded characters and appending to the global list
# params (html contents of page)
def extract_data(content):
    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
            GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

        lf.write(f"SUCCESSFUL     at {datetime.today()}\n")
        lf.write(f"setting up extraction objects     at {datetime.today()}\n")

    print(f"Setting up extraction objects...\n")

    # get all names and prices from html contents:
    all_names = content.find_all("div", class_="link")
    all_prices = content.find_all("div", class_="price")

    # make intermittent save lists:
    all_prices_list = []
    all_names_list = []

    # orders all prices in groups of four (buff, cheapest, market-cap, trade volume)
    all_prices_list_quad = []

    # strip of irrelevant characters and append every item in all_names and all_prices:
    # [0:-7] to strip of the "(Buy now)" text at the end of every name
    for i in all_names:
        all_names_list.append(i.text.translate(TRANSLATION_TABLE)[0:-7])

    for i in all_prices:
        all_prices_list.append(i.text.translate(TRANSLATION_TABLE))

    # pop the first item in all prices (entire cs:go market-cap)
    all_prices_list.pop(0)

    # pop all duplicate items in all_prices:
    for i in range((len(all_prices_list)) - 1, 0, -1):
        if abs(i) % 2 == 1:
            all_prices_list.pop(i)

    # group all elements by four(see up top):
    for i in range(0, len(all_prices_list) - 1, 4):
        all_prices_list_quad.append([all_prices_list[i], all_prices_list[(i + 1)],
                                     all_prices_list[(i + 2)], all_prices_list[(i + 3)]])

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
            GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:
        lf.write(f"Extracted data for current page:      at {datetime.today()}\n\n")
        lf.write(f"all names: {all_names_list}\n")
        lf.write(f"all prices: {all_prices_list}\n")
        lf.write(f"all prices in quad: {all_prices_list_quad}\n")

    # print all lists:
    print(f"all names:"
          f"{all_names_list}")
    print(f"all prices:"
          f"{all_prices_list}")
    print(f"all prices in quad"
          f"{all_prices_list_quad}\n")

    # check if page had loaded and items were added:
    # (always thirty elements per page, so len of name list should be 30)
    if len(all_names_list) <= 30:

        # print items to check.
        # print(f"checking1: {all_prices_list[0]}")
        # print(f"checking2: {all_names_list[0]}")

        print(f"saving to global save lists...")
        with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

            lf.write(f"saving to global save lists     at {datetime.today()}\n\n")
        # add all intermittent items to the global save lists:
        # names
        for i in all_names_list:
            all_names_save_list.append(i)

        # prices
        for k in all_prices_list:
            all_prices_save_list.append(k)

        # prices quad
        for j in all_prices_list_quad:
            all_prices_sub_save_list.append(j)

        try:
            if all_names_save_list[-1] == all_names_save_list[-31]:
                print(f"page reloaded same content, retrying")

                with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                        GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:
                    lf.write(f"Page reloaded same content, retrying     at {datetime.today()}... ")

                return False

        except IndexError:

            print(f"first list, so nothing to check")
            print("proceeding")


    # "throw error" if page doesn't load
    elif len(all_names_list) < 1:

        with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:
            lf.write(f"Page didn't load or empty, retrying     at {datetime.today()}... ")

        print("Page didnt load or empty, trying again")

        return False


# make spreadsheet with global save lists
# params (all names, all prices, calculate dmarket/steam arbitrage ?, sheet index)
def make_spreadsheet_normal(an, aps, cds, si):
    if GLOBAL_PARAM_DICT["exceptions"] is True:

        print(f"Exception when getting page contents..\n"
              f"Creating exception spreadsheet and writing to it...")

        possible_spreadsheets_list.append("exception.xlsx")

        exception_workbook = Workbook()
        exception_workbook.save(possible_spreadsheets_list[-1])

        si = -1

        with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

            lf.write(f"Exception when getting page contents..\n"
                     f"Creating exception spreadsheet and writing to it... "
                     f"proceeding      at {datetime.today()}\n")

    else:

        if GLOBAL_PARAM_DICT["daily_save"] is True:

            print(f"This is a daily save, creating new workbook\n"
                  f"Today is the {datetime.today()}")

            with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                    GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

                lf.write(f"This is a daily save ({DATE}), creating new workbook... "
                         f"proceeding      at {datetime.today()}\n")

            if ACTUAL_PARAM_LIST[1] == "y":

                possible_spreadsheets_list.append(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[0]daily_saves\\[0-1]only_cases\\daily_save_{str(DATE)}_onlycases.xlsx")

            else:

                possible_spreadsheets_list.append(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[0]daily_saves\\[0-2]normal\\daily_save_{str(DATE)}.xlsx")

            daily_workbook = Workbook()
            daily_workbook.save(possible_spreadsheets_list[-1])

            si = -1

        else:

            with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
                    GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

                lf.write(f"No exception when making workbook..\n"
                         f"proceeding as normal with spreadsheet creation... "
                         f"proceeding      at {datetime.today()}\n")

            print(f"No exception when making workbook..\n"
                  f"proceeding as normal with spreadsheet creation")

    print(f"Making spreadsheet...")

    # second translation tabel specific to this function to preserve some character
    translation_table_for_int = dict.fromkeys(map(ord, '.,$'), None)

    # load spreadsheet and format it
    workbook = load_workbook(str(possible_spreadsheets_list[si]))
    sheet = workbook.active
    sheet.title = "Main"
    sheet["A1"] = "Item name"
    sheet.column_dimensions["A"].width = 35
    sheet["B1"] = "Buff price"
    sheet.column_dimensions["B"].width = 25
    sheet["C1"] = "Cheapest price"
    sheet.column_dimensions["C"].width = 25
    sheet["D1"] = "Marketcap"
    sheet.column_dimensions["D"].width = 25
    sheet["E1"] = "Trade volume"
    sheet.column_dimensions["E"].width = 25
    sheet["F1"] = "estimated supply"
    sheet.column_dimensions["F"].width = 25
    sheet["G1"] = "gross estimated arbitrage"
    sheet.column_dimensions["G"].width = 25
    sheet["H1"] = "dmarket/steam arbitrage"
    sheet.column_dimensions["H"].width = 25

    row_start = 2  # start below the header row 2
    col_start = 1  # starts from column B

    print(f"adding data to spreadsheet...")

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
            GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

        lf.write(f"adding data rows to spreadsheet     at {datetime.today()}\n")

    # add all items into the rows:
    for i in range(0, len(an)):

        # add name at current index
        sheet.cell(row_start + i, col_start).value = an[i]

        # add buff price at current index
        sheet.cell(row_start + i, col_start + 1).value = aps[i][0]

        # add cheapest price at current index
        sheet.cell(row_start + i, col_start + 2).value = aps[i][1]

        # add market-cap at current index
        sheet.cell(row_start + i, col_start + 3).value = aps[i][2]

        # add trade volume at current index
        sheet.cell(row_start + i, col_start + 4).value = aps[i][3]

        # add estimated supply at current index
        sheet.cell(row_start + i, col_start + 5).value = str(float(aps[i][2].translate(translation_table_for_int))

                                                             / float(

            aps[i][1].translate(translation_table_for_int)))

        # add cheapest/buff arbitrage at current index
        sheet.cell(row_start + i, col_start + 6).value = str(round(int(aps[i][1].translate(translation_table_for_int))

                                                                   / int(

            aps[i][0].translate(translation_table_for_int)), 2)) + "%"

        # add dmarket/steam arbitrage at current index if requested
        if cds is True:

            sheet.cell(row_start + i, col_start + 7).value = calculate_dmarket_steam_arbitrage(sn=an[i], fn=False,
                                                                                               sf="")

        else:
            pass

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
            GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

        lf.write(f"saving spreadsheet (name: {possible_spreadsheets_list[si]})...\n"
                 f"saved {len(all_names_save_list)} rows successfully to {possible_spreadsheets_list[si]}      at {datetime.today()}\n")

    print(f"saving spreadsheet {possible_spreadsheets_list[si]}...")
    print(f"saved {len(all_names_save_list)} rows successfully to {possible_spreadsheets_list[si]}")

    # save spreadsheet
    workbook.save(str(possible_spreadsheets_list[si]))

    return si


# function to print results and possible buy options to a text file

def print_results(si):
    # second translation tabel specific to this function to preserve some character
    translation_table_for_print = dict.fromkeys(map(ord, '%'), None)

    results_numbers_list = []
    collumns_list = ["A", "B", "C", "D", "E", "F", "G"]

    wb = load_workbook(str(possible_spreadsheets_list[int(si)]))
    sheet = wb.active

    print(f"gathering results...")

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
            GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

        lf.write(f"gathering results to proceed with result print (file: ")

    for i in range(2, len(all_names_save_list)):

        if float(str(sheet[f"G{str(i)}"].value).translate(translation_table_for_print)) < 0.7:

            results_numbers_list.append(i)

        else:

            pass

    file_index = 0
    i = 0
    while not i == 20:

        file_index += 1

        # print(f"File index is {file_index}")
        # print(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[1]result_files\\results_for_{str(DATE)}[{str(file_index)}].txt")
        # print("FILE_exists?: " + str(os.path.exists(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[1]result_files\\results_for_{str(DATE)}[{str(file_index)}].txt")))

        if os.path.exists(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[1]result_files\\results_for_{str(DATE)}[{str(file_index)}].txt") is True:

            pass

        elif os.path.exists(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[1]result_files\\results_for_{str(DATE)}[{str(file_index)}].txt") is False:

            break

        elif i == 20:

            print(f"Aborting save to txt file. There can't be more than 20.")

        i += 1

    print(f"writing results to text file ")

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
            GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

        lf.write(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[1]result_files\\results_for_{str(DATE)}[{str(file_index)}].txt)       at {datetime.today()}\n")

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[1]result_files\\results_for_{str(DATE)}[{str(file_index)}].txt", "x") as rf:

        rf.write(f"Results file for search on {str(DATETIME)} with search params:\n\n")

        for i in range(0, len(PARAMS_LIST) - 1):
            rf.write(f"{str(PARAMS_LIST[i])} --- {str(ACTUAL_PARAM_LIST[i])}\n")

        rf.write("\n")
        curr_option_index = 1

        for i in results_numbers_list:

            helper_result_list = []

            for k in collumns_list:
                helper_result_list.append(str(sheet[f"{str(k)}{str(i)}"].value))

            print(f"Option {str(curr_option_index)}: {str(helper_result_list)}")

            rf.write(f"Option {str(curr_option_index)}: {str(helper_result_list[0])}:\n"
                     f" -buff price: {helper_result_list[1]}\n"
                     f" -cheapest price: {helper_result_list[2]}\n"
                     f" -market-cap: {helper_result_list[3]}\n"
                     f" -trade volume: {helper_result_list[4]}\n"
                     f" -estimated supply: {helper_result_list[5]}\n"
                     f" -gross estimated arbitrage: {helper_result_list[6]}\n"
                     f"\n")

            curr_option_index += 1

    with open(f"C:\\Users\\felix_a0jy\\PycharmProjects\\pricempire_scraper\\[2]log_files\\log_file_for_{str(DATE)}[" + str(
            GLOBAL_PARAM_DICT["log_file_index"]) + f"].txt", "a", encoding="utf-8") as lf:

        lf.write(f"Succesfully wrote to result file     at {datetime.today()}\n")


# possible function to use selenium with proxy:
def get_proxy():
    pass

    # proxy_dict = {"20.210.113.32": "80", "71.86.129.131": "8080", "34.23.45.223": "80", "43.255.113.232": "80",
    #              "20.24.43.214": "80"}


# function to calculate dmarket/steam arbitrage if requested
# params (skin name, family_needed(if name is similar to others), skin_famaily(NULL by default)
def calculate_dmarket_steam_arbitrage(sn, fn, sf):
    # get dmarket cheapest price and wait
    dm_price = get_dmarket_html(sn, fn, sf)
    time.sleep(1)

    # get steam price
    st_price = get_steam_html(sn)
    time.sleep(1)

    # calculate arbitrage
    dmst_arbitrage = dm_price / float(st_price)

    # return value to spreadsheet
    return dmst_arbitrage


# function to get dmarket cheapest price:
# params (skin name, family_needed(if name is similar to others), skin_famaily(NULL by default)
def get_dmarket_html(skin_name, family_needed, skin_family):
    # check if you need the family of the skin
    if family_needed is True:

        # create dynamic url
        DMARKET_URL = f"https://dmarket.com/de/ingame-items/item-list/csgo-skins?family={skin_family}&title={skin_name}"

    else:

        # create dynamic url
        DMARKET_URL = f"https://dmarket.com/de/ingame-items/item-list/csgo-skins?title={skin_name}"

    # driver (currently Edge)
    driver = webdriver.Edge()

    # go to site and wait for page load
    driver.get(DMARKET_URL)
    time.sleep(5)

    # click buttons to filter by lowest price and wait for page load
    driver.find_element(By.CSS_SELECTOR, "div[class='o-select__sortTexts'] span").click()
    driver.find_element(By.CSS_SELECTOR, "button:nth-child(6) strong:nth-child(1)").click()
    time.sleep(WAIT_TIME)

    # get html_contents
    contents = BeautifulSoup(driver.page_source, "html.parser")

    # start function which will return the extracted price value
    return extract_dmarket_contents(contents)


# function to extract the price value out of the html contents
# params (html contents)
def extract_dmarket_contents(cts):
    # third translation tabel specific to this function to preserve some character
    translation_table = dict.fromkeys(map(ord, '$'), None)

    # extract all price, quality and float values from html contents
    # all_names = currently cant extract names
    all_prices = cts.find_all("price", class_="ng-star-inserted")
    all_qualities = cts.find_all("a", class_="c-asset__exterior c-asset__exterior--link ng-star-inserted")
    all_floats = cts.find_all("span", class_="o-blur")

    # create all intermittent lists
    # all_names_list = []

    all_prices_list = []
    all_qualities_list = []
    all_floats_list = []

    # strip and append values to intermittent lists
    # all prices
    for i in all_prices:
        all_prices_list.append(i.text.translate(translation_table))

    # all qualities
    for i in all_qualities:
        all_qualities_list.append(i.text)

    # all floats
    for i in all_floats:
        all_floats_list.append(i.text)

    # print all lists
    print(f"all prices(dmarket)"  # ignore the \xa (it's a comma)
          f"{all_prices_list}")

    print(f"all qualities"
          f"{all_qualities_list}")

    print(f"all floats"
          f"{all_floats_list}")

    # try to return a number, if it throws error do some formatting
    try:

        # return as float
        return float(all_prices_list[0])

    # catch ValueError
    except ValueError:

        # some fucked up shit thats not working
        new_int = str(all_prices_list[0][0]) + str(all_prices_list[0][2:-1])
        new_new_float = ""
        help_float_list = []

        for i in new_int:
            help_float_list.append(i)

        for k in help_float_list:
            new_new_float = new_new_float + k

        return float(new_new_float)


# function to get steam html
# params (skin name)
def get_steam_html(skin_name):
    # steam url
    STEAM_URL = f"https://steamcommunity.com/market/search?appid=730&q={skin_name}"

    # driver (currently Edge)
    driver = webdriver.Edge()

    # go to site and wait for page load
    driver.get(STEAM_URL)
    time.sleep(WAIT_TIME)

    # get html contents
    contents = BeautifulSoup(driver.page_source, "html.parser")

    # call function which will return the cheapest steam price
    return extract_steam_data(contents)


# function which extracts steam html contents
# params (html contents of page)
def extract_steam_data(cts):
    # third translation tabel specific to this function to preserve some character
    translation_table = dict.fromkeys(map(ord, '$USD\n\tarting: '), None)

    # extract all name and price values on page
    all_names = cts.find_all("span", class_="market_listing_item_name")
    all_prices = cts.find_all("span", class_="normal_price")

    # create all intermittent lists
    all_steam_names_list = []
    all_steam_prices_list = []

    # strip and append all items to intermittent lists
    for i in all_names:
        all_steam_names_list.append(i.text)

    for k in all_prices:
        all_steam_prices_list.append(k.text.translate(translation_table))

    # pop all duplicate price items
    for i in range(len(all_steam_prices_list) - 1, -1, -1):

        if abs(i) % 2 == 0:

            all_steam_prices_list.pop(i)
        else:
            pass

    # print all lists
    print(f"all names (steam)"
          f"{all_steam_names_list}")
    print(f"all prices (steam)"
          f"{all_steam_prices_list}")

    # return cheapest price
    return float(all_steam_prices_list[0])


# start main function for testing

main(LOG_FILE_INDEX)

# calculate_dmarket_steam_arbitrage("AWP | Asiimov (Field-Tested)", False, "")

# if __name__ == "main":
#         main   ()
