# IMPORTS...................................
from bs4 import BeautifulSoup
import selenium.common.exceptions
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from datetime import datetime, date
import os
import json

# "Global" objects...........................

# global save lists:
ALL_NAMES_SAVE_LIST = []
ALL_PRICES_SAVE_LIST = []
ALL_PRICES_QUAD_SAVE_LIST = []

# list with premade file names (expandable)
possible_spreadsheets_list = ["C:\\Users\\Public\\pricempire_web_scraper\\[3]any_run_saves\\normal_run_save_file.xlsx",
                              "C:\\Users\\Public\\pricempire_web_scraper\\[3]any_run_saves\\onlycases_run_save_file.xlsx"]

# dictionary's and lists for global item accessibility
GLOBAL_PARAM_DICT = {"website_url": "", "no_content_exception": False, "daily_save": False, "log_file_index": "", "result_file_index": "", "sheet_index": 0}
PARAMS_DICT = {}
DEFAULT_PARAMS_DICT = {}

# global values
WAIT_TIME = 10

# default translation table
TRANSLATION_TABLE = dict.fromkeys(map(ord, '\n\t'), None)

# datetime at the start of the program
DATETIME = datetime.today()
DATE = date.today()

# paths
LOG_PATH = "C:\\Users\\Public\\pricempire_web_scraper\\[2]log_files\\log_file_for_"
RESULTS_PATH = "C:\\Users\\Public\\pricempire_web_scraper\\[1]result_files\\results_for_"


# FUNCTIONS...................................

# main function:
def main_func():
    # get log file index
    get_file_index(file_path=str(LOG_PATH), global_param_name="log_file_index")

    # get result file index
    get_file_index(file_path=str(RESULTS_PATH), global_param_name="result_file_index")

    write_to_file(file_type="log", content=f"Log file " + str(GLOBAL_PARAM_DICT["log_file_index"]) + f" for normal run\n"
                                                                                                     f"at {datetime.today()}|: Starting log\n")

    # gets defaults params
    get_default_params()

    # offer explanation
    starting_explanation()

    # collect all params
    get_params()

    # assigns website url
    GLOBAL_PARAM_DICT["website_url"] = "https://pricempire.com/trending?to={}&sort={}{}&blacklist={}&search={}".format(int(PARAMS_DICT["limit"]),
                                                                                                                       str(PARAMS_DICT["sort"]),
                                                                                                                       str(PARAMS_DICT["order"]),
                                                                                                                       str(PARAMS_DICT["blacklist"]),
                                                                                                                       str(PARAMS_DICT["search"]))
    # catches errors to put in the log file
    try:

        # functions:

        # get the html contents
        get_html()

        # makes spreadsheet
        make_spreadsheet_normal()

        # assigns the print results parameter
        print_results_yn = input("Print result text file? (y/n) ").lower()  # decides if program should make a spreadsheet with all the data
        while print_results_yn not in ("", "y", "n"):
            print("Please enter a valid parameter: (y/n) or just hit enter!")
            print_results_yn = input("Print result text file? (y/n) ").lower()

        # prints results if requested
        if print_results_yn == "y":
            print_results()

        # adds log text 3 (program success (if succeeded)
        write_to_file(file_type="log", content=f"at {datetime.today()}|: completed program without errors\n")

    except Exception as e:

        print(f"Error when running program")

        # adds log text 5 (program unsuccessful(if unsuccessful))
        write_to_file(file_type="log", content=f"at {datetime.today()}|: did not complete program regularly\n")
        write_to_file(file_type="log", content=f"error was: {str(e)}\n")

        print(e)


# get html contents of page and start extraction function
def get_html():

    # get website url
    accessed_website = GLOBAL_PARAM_DICT["website_url"]

    # driver (currently Edge):
    driver = webdriver.Edge()

    # go to site and wait for first page load
    driver.get(accessed_website)
    print(f"{accessed_website}\n...")
    time.sleep(5)

    # try to click the cookies consent button if it exists
    try:

        driver.find_element(By.CSS_SELECTOR, ".fc-cta-consent").click()

    except selenium.common.exceptions.NoSuchElementException:

        print("Button not found")

    # adds to log 5 (accessed website)
    write_to_file(file_type="log", content=f"at {datetime.today()}|: accessed website ({accessed_website})\n")

    # page load time:
    time.sleep(WAIT_TIME)

    # get html contents of all pages by clicking next page button and starting extraction function
    for i in range(0, int(PARAMS_DICT["iterations"])):

        print(f"Getting Page ({i + 1}) contents...")

        # adds to log 6 (confirmation for getting page contents)
        write_to_file(file_type="log", content=f"at {datetime.today()}|: Getting Page ({i + 1}) contents... ")

        # get the contents and start the extraction function
        try:

            # get contents:
            contents = BeautifulSoup(driver.page_source, "html.parser")

            print(f"extracting data...")

            # try to extract the data into a list
            if extract_data(contents) is False:  # if fails, reload page then try again

                # adds to log 7 (unsuccessful extraction)
                write_to_file(file_type="log", content=f"UNSUCCESSFUL\n")
                write_to_file(file_type="log", content=f"at {datetime.today()}|: attempting to reload page... ")

                # reload page and wait for page load longer this time:
                driver.refresh()
                time.sleep(WAIT_TIME + 10)

                # get contents:
                contents_reload = BeautifulSoup(driver.page_source, "html.parser")

                print(f"extracting data...")

                # try again:
                if extract_data(contents_reload) is False:  # try to access next page by link

                    # adds to log 8 (unsuccessful extraction)
                    write_to_file(file_type="log", content=f"UNSUCCESSFUL\n")
                    write_to_file(file_type="log", content=f"at {datetime.today()}|: attempting to skip to next page... ")

                    print(f"attempting to skip to next page... ")

                    # close and reopen driver go to page by url and wait for page load

                    driver.quit()
                    time.sleep(60)

                    # maybe add to try other browser
                    driver = webdriver.Edge()
                    driver.get(str(GLOBAL_PARAM_DICT["website_url"]) + str(f"&page={str(i)}"))
                    time.sleep(WAIT_TIME + 10)

                    # get contents
                    contents_reload_2 = BeautifulSoup(driver.page_source, "html.parser")

                    if extract_data(contents_reload_2) is False:
                        write_to_file(file_type="log", content=f"UNSUCCESSFUL\n")
                        write_to_file(file_type="log", content=f"at {datetime.today()}|: aborting and saving current items (amount: {str(len(ALL_NAMES_SAVE_LIST))})\n")

                        print(f"Something broke, saving current items (amount: {str(len(ALL_NAMES_SAVE_LIST))}) and aborting")

                        GLOBAL_PARAM_DICT["no_content_exception"] = True

                        break

            else:

                pass

            # click the next page button and wait for the page to load then repeat
            try:

                driver.find_element(By.CSS_SELECTOR, ".fas.fa-chevron-right").click()
                time.sleep(WAIT_TIME)

            except selenium.common.exceptions.NoSuchElementException:

                print(f"No next page button so probably last page")

                break

        except Exception as e:

            print(f"Trying to catch all possible errors here")
            print(str(e))


# extract the data by stripping all unneeded characters and appending to the global list
# params (html contents of page)
def extract_data(content):

    write_to_file(file_type="log", content=f"SUCCESSFUL\n")
    write_to_file(file_type="log", content=f"at {datetime.today()}|: setting up extraction objects\n")

    print(f"Setting up extraction objects...\n")

    # get all names and prices from html contents:
    all_names = content.find_all("div", class_="link")
    all_prices = content.find_all("div", class_="price")

    # make intermittent save lists:
    all_prices_list = []
    all_names_list = []

    # orders all prices in groups of four (buff, cheapest, market-cap, trade volume)
    all_prices_list_quad = []

    # strip of irrelevant characters and append every item in all_names and all_prices:
    # [0:-7] to strip of the "(Buy now)" text at the end of every name
    for i in all_names:
        all_names_list.append(i.text.translate(TRANSLATION_TABLE)[0:-7])

    for i in all_prices:
        all_prices_list.append(i.text.translate(TRANSLATION_TABLE))

    # pop the first item in all prices (entire cs:go market-cap)
    all_prices_list.pop(0)

    # pop all duplicate items in all_prices:
    for i in range((len(all_prices_list)) - 1, 0, -1):
        if abs(i) % 2 == 1:
            all_prices_list.pop(i)

    # group all elements by four(see up top):
    for i in range(0, len(all_prices_list) - 1, 4):
        all_prices_list_quad.append([all_prices_list[i], all_prices_list[(i + 1)],
                                     all_prices_list[(i + 2)], all_prices_list[(i + 3)]])

    write_to_file(file_type="log", content=f"at {datetime.today()}|: Extracted data for current page:\n\n")
    write_to_file(file_type="log", content=f"all names: {all_names_list}\n"
                                           f"all prices: {all_prices_list}\n"
                                           f"all prices in quad: {all_prices_list_quad}\n")

    # print all lists:
    print(f"all names:"
          f"{all_names_list}")
    print(f"all prices:"
          f"{all_prices_list}")
    print(f"all prices in quad"
          f"{all_prices_list_quad}\n")

    # check if page had loaded and items were added:
    # (always thirty elements per page, so len of name list should be 30)
    if 30 >= len(all_names_list) > 0:

        # print items to check.
        # print(f"checking1: {all_prices_list[0]}")
        # print(f"checking2: {all_names_list[0]}")

        print(f"saving to global save lists...")

        write_to_file(file_type="log", content=f"at {datetime.today()}|: saving to global save lists\n\n")

        # names
        for i in all_names_list:
            ALL_NAMES_SAVE_LIST.append(i)

        # prices
        for k in all_prices_list:
            ALL_PRICES_SAVE_LIST.append(k)

        # prices quad
        for j in all_prices_list_quad:
            ALL_PRICES_QUAD_SAVE_LIST.append(j)

        try:

            if ALL_NAMES_SAVE_LIST[-1] == ALL_NAMES_SAVE_LIST[-31]:
                print(f"page reloaded same content, retrying")

                write_to_file(file_type="log", content=f"at {datetime.today()}|: Page reloaded same content, retrying... ")

                return False

        except IndexError:

            print(f"first page, so nothing to check")
            print("proceeding")

    # "throw error" if page doesn't load
    elif len(all_names_list) < 1:

        write_to_file(file_type="log", content=f"at {datetime.today()}|: Page didn't load or empty, retrying... ")

        print("Page didn't load or empty, trying again")

        return False


# make spreadsheet with global save lists
# params (all names, all prices, calculate dmarket/steam arbitrage ?, sheet index)
def make_spreadsheet_normal():

    if GLOBAL_PARAM_DICT["no_content_exception"] is True:

        print(f"Exception when getting page contents..\n"
              f"Creating exception spreadsheet and writing to it...")

        possible_spreadsheets_list.append("C:\\Users\\Public\\pricempire_web_scraper\\[4]exception_run_saves\\exception_run_save.xlsx")

        exception_workbook = Workbook()
        exception_workbook.save(possible_spreadsheets_list[-1])

        # assign proper sheet index
        GLOBAL_PARAM_DICT["sheet_index"] = -1

        write_to_file(file_type="log", content=f"at {datetime.today()}|: Exception when getting page contents..\n"
                                               f"at {datetime.today()}|: Creating exception spreadsheet and writing to it... "
                                               f"proceeding\n")

    else:

        if GLOBAL_PARAM_DICT["daily_save"] is True:

            print(f"This is a daily save, creating new workbook\n"
                  f"Today is the {datetime.today()}")

            write_to_file(file_type="log", content=f"at {datetime.today()}|: This is a daily save ({DATE}), creating new workbook... "
                                                   f"proceeding\n")

            if str(PARAMS_DICT["only_cases"]) == "y":

                possible_spreadsheets_list.append(f"C:\\Users\\Public\\pricempire_web_scraper\\[0]daily_saves\\[0-1]only_cases\\daily_save_{str(DATE)}_onlycases.xlsx")

            else:

                possible_spreadsheets_list.append(f"C:\\Users\\Public\\pricempire_web_scraper\\[0]daily_saves\\[0-2]normal\\daily_save_{str(DATE)}.xlsx")

            daily_workbook = Workbook()
            daily_workbook.save(possible_spreadsheets_list[-1])

            write_to_file(file_type="log", content=f"at {datetime.today()}|: created and saved workbook successfully (name: {str(possible_spreadsheets_list[-1])})\n")

            GLOBAL_PARAM_DICT["sheet_index"] = -1

        else:

            write_to_file(file_type="log", content=f"at {datetime.today()}|: No exception when making workbook..\n"
                                                   f"at {datetime.today()}|: proceeding as normal with spreadsheet creation... "
                                                   f"proceeding\n")

            print(f"No exception when making workbook..\n"
                  f"proceeding as normal with spreadsheet creation")

    print(f"Making spreadsheet...")

    # second translation tabel specific to this function to preserve some character
    translation_table_for_int = dict.fromkeys(map(ord, '.,$'), None)

    # load spreadsheet and format it
    workbook = load_workbook(str(possible_spreadsheets_list[int(GLOBAL_PARAM_DICT["sheet_index"])]))
    sheet = workbook.active
    sheet.title = "Main"
    sheet["A1"] = "Item name"
    sheet.column_dimensions["A"].width = 35
    sheet["B1"] = "Buff price"
    sheet.column_dimensions["B"].width = 25
    sheet["C1"] = "Cheapest price"
    sheet.column_dimensions["C"].width = 25
    sheet["D1"] = "Marketcap"
    sheet.column_dimensions["D"].width = 25
    sheet["E1"] = "Trade volume"
    sheet.column_dimensions["E"].width = 25
    sheet["F1"] = "estimated supply"
    sheet.column_dimensions["F"].width = 25
    sheet["G1"] = "gross estimated arbitrage"
    sheet.column_dimensions["G"].width = 25
    sheet["H1"] = "dmarket/steam arbitrage"
    sheet.column_dimensions["H"].width = 25

    row_start = 2  # start below the header row 2
    col_start = 1  # starts from column B

    print(f"adding data to spreadsheet...")

    write_to_file(file_type="log", content=f"at {datetime.today()}|: adding data rows to spreadsheet\n")

    # add all items into the rows:
    for i in range(0, len(ALL_NAMES_SAVE_LIST)):

        # add name at current index
        sheet.cell(row_start + i, col_start).value = ALL_NAMES_SAVE_LIST[i]

        # add buff price at current index
        sheet.cell(row_start + i, col_start + 1).value = ALL_PRICES_QUAD_SAVE_LIST[i][0]

        # add cheapest price at current index
        sheet.cell(row_start + i, col_start + 2).value = ALL_PRICES_QUAD_SAVE_LIST[i][1]

        # add market-cap at current index
        sheet.cell(row_start + i, col_start + 3).value = ALL_PRICES_QUAD_SAVE_LIST[i][2]

        # add trade volume at current index
        sheet.cell(row_start + i, col_start + 4).value = ALL_PRICES_QUAD_SAVE_LIST[i][3]

        # add estimated supply at current index
        sheet.cell(row_start + i, col_start + 5).value = str(
            float(ALL_PRICES_QUAD_SAVE_LIST[i][2].translate(translation_table_for_int)) / float(ALL_PRICES_QUAD_SAVE_LIST[i][1].translate(translation_table_for_int)))

        # add cheapest/buff arbitrage at current index
        sheet.cell(row_start + i, col_start + 6).value = str(
            round(float(ALL_PRICES_QUAD_SAVE_LIST[i][1].translate(translation_table_for_int)) / float(ALL_PRICES_QUAD_SAVE_LIST[i][0].translate(translation_table_for_int)), 2)) + "%"

        # add dmarket/steam arbitrage at current index if requested
        if PARAMS_DICT["calc_dmarket_steam"] == "y":

            sheet.cell(row_start + i, col_start + 7).value = calculate_dmarket_steam_arbitrage(sn=ALL_NAMES_SAVE_LIST[i], fn=False,
                                                                                               sf="")

        else:
            pass

    si = int(GLOBAL_PARAM_DICT["sheet_index"])  # current solution

    write_to_file(file_type="log", content=f"at {datetime.today()}|: saving spreadsheet (name: {possible_spreadsheets_list[si]})...\n"
                                           f"at {datetime.today()}|: saved {len(ALL_NAMES_SAVE_LIST)} rows ({len(ALL_NAMES_SAVE_LIST)*6} data points) successfully to {possible_spreadsheets_list[si]}")

    print(f"saving spreadsheet {possible_spreadsheets_list[si]}...")
    print(f"saved {len(ALL_NAMES_SAVE_LIST)} rows successfully to {possible_spreadsheets_list[si]}")

    # save spreadsheet
    workbook.save(str(possible_spreadsheets_list[si]))


# function to print results and possible buy options to a text file
def print_results():

    # second translation tabel specific to this function to preserve some character
    translation_table_for_print = dict.fromkeys(map(ord, '%'), None)

    results_numbers_list = []
    columns_list = ["A", "B", "C", "D", "E", "F", "G"]

    wb = load_workbook(str(possible_spreadsheets_list[int(GLOBAL_PARAM_DICT["sheet_index"])]))
    sheet = wb.active

    print(f"gathering results...")

    write_to_file(file_type="log", content=f"gathering results to proceed with result print (file: ")

    for i in range(2, len(ALL_NAMES_SAVE_LIST)):

        if float(str(sheet[f"G{str(i)}"].value).translate(translation_table_for_print)) < PARAMS_DICT["min_arbitrage"]:  # min arbitrage

            results_numbers_list.append(i)

        else:

            pass

    write_to_file(file_type="log", content=f"at {datetime.today()}|: created result file (name: C:\\Users\\Public\\pricempire_web_scraper\\[1]result_files\\results_for_{str(DATE)}[" + str(
        GLOBAL_PARAM_DICT["result_file_index"]) + f"].txt)\n")

    with open(f"C:\\Users\\Public\\pricempire_web_scraper\\[1]result_files\\results_for_{str(DATE)}[" + str(GLOBAL_PARAM_DICT["result_file_index"]) + f"].txt", "x", encoding="utf-8") as rf:

        rf.write(f"Results file for search on {str(DATETIME)} with search params:\n\n")

        for i in PARAMS_DICT:
            rf.write(f"{str(i)} --- {str(PARAMS_DICT[i])}\n")

        rf.write("\n")
        curr_option_index = 1

        for i in results_numbers_list:

            helper_result_list = []

            for k in columns_list:
                helper_result_list.append(str(sheet[f"{str(k)}{str(i)}"].value))

            print(f"Option {str(curr_option_index)}: {str(helper_result_list)}")

            rf.write(f"Option {str(curr_option_index)}: {str(helper_result_list[0])}:\n"
                     f" -buff price: {helper_result_list[1]}\n"
                     f" -cheapest price: {helper_result_list[2]}\n"
                     f" -market-cap: {helper_result_list[3]}\n"
                     f" -trade volume: {helper_result_list[4]}\n"
                     f" -estimated supply: {helper_result_list[5]}\n"
                     f" -gross estimated arbitrage: {helper_result_list[6]}\n"
                     f"\n")

            curr_option_index += 1

        if len(results_numbers_list) == 0:
            rf.write(f"No options available right now\n")

    write_to_file(file_type="log", content=f"at {datetime.today()}|: Successfully wrote to result file\n")


def starting_explanation():

    print("PRICEMPIRE-WEB-SCRAPER by Felix Zabudkin \n")

    show_explanation = input("Do you want to see a explanation for each parameter? (y/n) ").lower()
    while show_explanation not in ("", "y", "n"):
        print("Please enter a valid parameter: (y/n) or just hit enter!")
        show_explanation = input("Do you want to see a explanation for each parameter? (y/n) ").lower()

    if show_explanation == "y":

        print(f"\n"
              f"1. only cases --- decides if search should only be for cases (options: y/n) \n"
              f"2. daily save --- decides if search should be saved as a daily save (options: y/n) \n"
              f"3. limit --- decides what the price limit should be (options: any positive integer) \n"
              f"4. sort --- decides how the items should be sorted (options: buff, cheapest, marketcap, tradevolume) \n"
              f"5. order --- decides if items should be sorted ascending or descending (options: y/n) \n"
              f"6. blacklist --- decides which items or chars should be blacklisted from the search (options: any char or string) NOTE: this parameter will not always be shown \n"
              f"7. search --- decides if program should search for a specific string or char (options: any char or string)  NOTE: this parameter will not always be shown \n"
              f"8. cds --- decides if program should add the dmarket/steam market-price-difference respectively arbitrage to the save file (options: y/n) NOTE: this feature is still in work, so a y won't do anything \n"
              f"9. print results -- decides if a result file with buy options should be made (options: y/n) \n"
              f"10. min arbitrage --- decides what the minimum arbitrage for the result file filter should be (options: any positive float between 0 and 1)\n"
              f"\n"
              f"You have the option to just press enter for every parameter and the respective value will be the default. \n")

    show_default_param_values = input("Show default parameter values? (y/n) ").lower()
    while show_default_param_values not in ("", "y", "n"):
        print("Please enter a valid parameter: (y/n) or just hit enter!")
        show_default_param_values = input("Show default parameter values? (y/n) ").lower()

    if show_default_param_values == "y":
        print(f"\ndefault for...")
        for i in DEFAULT_PARAMS_DICT:
            print(f"    {str(i)} -> {str(DEFAULT_PARAMS_DICT[str(i)])} ")

        start_it = input("\nPress Enter to continue")

    print("Starting program... \n\n")
    print("Please select your parameters.")


def get_file_index(file_path, global_param_name):

    # creating error list index
    i = 0
    lfi = 0

    while not i == 21:  # 21 because there can be only 20 log files per day. Function searches which log file names already exist and takes one which doesn'

        if os.path.exists(f"{str(file_path)}{str(DATE)}[{str(lfi)}].txt") is False:

            break

        lfi += 1
        i += 1

    # saves the log file index of the current run
    GLOBAL_PARAM_DICT[str(global_param_name)] = int(lfi)


def get_params():

    # parameters:
    only_cases = input("Search only for cases? (y/n) ").lower()  # decides if to only search for cases or not
    while only_cases not in ("", "y", "n"):
        print("Please enter a valid parameter: (y/n) or just hit enter!")
        only_cases = input("Search only for cases? (y/n) ").lower()

    daily_save = input("Is this a daily save for historical data? (y/n) ").lower()  # decides if run is treated as daily safe
    while daily_save not in ("", "y", "n"):
        print("Please enter a valid parameter: (y/n) or just hit enter!")
        daily_save = input("Is this a daily save for historical data? (y/n) ").lower()

    limit = input("Price limit? (positive integer) ").lower()  # decides the price limit of the search
    while is_any(needed_type="int", value=limit) is False and not limit == "":
        print("Please enter a valid parameter: (ANY POSITIVE INTEGER) or just hit enter!")
        limit = input("Price limit? (integer) ").lower()

    sort = input("Sort by ...? (buff, cheapest, marketcap, tradevolume) ").lower()  # decides how the search items are sorted
    while sort not in ("", "buff", "cheapest", "marketcap", "tradevolume"):
        print("Please enter a valid parameter: (buff, cheapest, marketcap, tradevolume) or just hit enter!")
        sort = input("Sort by ...? (buff, cheapest, marketcap, tradevolume) ").lower()

    order = input("Sort ascending or descending? (a/d) ").lower()  # decides if search items are ASC or DESC
    while order not in ("", "a", "d"):
        print("Please enter a valid parameter: (a/d) or just hit enter!")
        order = input("Sort ascending or descending? (a/d) ").lower()

    if only_cases == "" or only_cases == "n":

        blacklist = input("Blacklist items: (comma separated, no space) ").lower()  # blacklists characters

    elif only_cases == "y":

        blacklist = "hardened"

    else:

        blacklist = ""

    if only_cases == "" or only_cases == "n":

        search = input("Search for (a) specific item/s: ").lower()  # takes specific search arguments

    elif only_cases == "y":

        search = "case"

    else:

        search = ""

    if only_cases == "" or only_cases == "n":

        if daily_save == "y":

            iterations = ""

        else:

            iterations = input("How many pages to scrape? (positive integer) ")  # decides how many pages will be scraped
            while is_any(needed_type="int", value=iterations) is False and not iterations == "":
                print("Please enter a valid parameter: (ANY  POSITIVE INTEGER) or just hit enter!")
                iterations = input("How many pages to scrape? (positive integer) ")
    else:

        iterations = ""

    min_arbitrage = input("Minimum arbitrage for result file? (float between 0 and 1) ")  # decides if to only search for cases or not
    while is_any(needed_type="float01", value=min_arbitrage) is False and not min_arbitrage == "":
        print("Please enter a valid parameter: (float between 0 and 1) or just hit enter!")
        min_arbitrage = input("Minimum arbitrage for result file? (float between 0 and 1) ")

    calc_dmarket_steam = input("Calculate dmarket/steam arbitrage? Currently not working! (y/n) ").lower()  # decides if the dmarket/steam arbitrage should be calculated for each item
    while calc_dmarket_steam not in ("", "y", "n"):
        print("Please enter a valid parameter: (y/n) or just hit enter!")
        calc_dmarket_steam = input("Calculate dmarket/steam arbitrage? Currently not working! (y/n) ").lower()

    # default params:

    # limit:
    if limit == "":
        limit = DEFAULT_PARAMS_DICT["limit"]

    # sort
    if sort == "":
        sort = DEFAULT_PARAMS_DICT["sort"]

    # order
    if order == "" or order == "d":
        order = DEFAULT_PARAMS_DICT["order"]

    elif order == "a":
        order = ":ASC"

    # only cases
    if only_cases == "":
        only_cases = DEFAULT_PARAMS_DICT["only_cases"]

    # iterations
    if iterations == "":

        if only_cases == "y":

            iterations = 2

        elif daily_save == "y":

            iterations = 715

        else:

            iterations = DEFAULT_PARAMS_DICT["iterations"]

    # calc dmarket/steam arbitrage
    if calc_dmarket_steam == "":
        calc_dmarket_steam = DEFAULT_PARAMS_DICT["calc_dmarket_steam"]

    elif calc_dmarket_steam == "y":

        calc_dmarket_steam = "n"  # because it currently doesn't work

    # daily save
    if daily_save == "":

        daily_save = DEFAULT_PARAMS_DICT["daily_save"]

    elif daily_save == "y":

        GLOBAL_PARAM_DICT["daily_save"] = True

    # assign sheet index:
    if only_cases == "y":

        GLOBAL_PARAM_DICT["sheet_index"] = 1

    else:

        # here more options can be added
        GLOBAL_PARAM_DICT["sheet_index"] = 0

    # print all info
    print("assigned params are:")
    for i in PARAMS_DICT:
        print(f"    {str(i)} -> {str(PARAMS_DICT[i])}")

    print(
        f"\n...\nEverything Ok\n"
        f"Expected run-time: {round(float(((int(iterations) * 10) + 40) / 60), 1)} min. (+ 30-120 sec if errors occur)\n"
        f"Expected amount of items: {int(int(iterations) * 30 * 4)}\n"
        f"...\n"
        f"Getting html contents...\n")

    # append all params to the list which saves them
    PARAMS_DICT.update({"daily_save" : daily_save,
                        "only_cases" : only_cases,
                        "limit" : limit,
                        "sort" : sort,
                        "order" : order,
                        "blacklist" : blacklist,
                        "search": search, "iterations" : iterations,
                        "calc_dmarket_steam" : calc_dmarket_steam,
                        "min_arbitrage" : min_arbitrage})

    # adds log text 2 (all params)
    write_to_file(file_type="log", content=f"at {datetime.today()}|: Params assigned\n\n")

    for i in PARAMS_DICT:
        write_to_file(file_type="log", content=f"{str(i)} --- {str(PARAMS_DICT[i])}\n")
    write_to_file(file_type="log", content="\n")


# function whích writes to log or result file
def write_to_file(file_type, content):

    if file_type == "log":

        file_path = LOG_PATH

    elif file_type == "result":

        file_path = RESULTS_PATH

    with open(f"{str(file_path)}{str(DATE)}[" + str(GLOBAL_PARAM_DICT["log_file_index"]) + "].txt", "a", encoding="utf-8") as lf:

        lf.write(str(content))


def get_default_params():

    with open("C:\\Users\\Public\\pricempire_web_scraper\\config\\config.json", "r") as cf:

        defaults = json.load(cf)['defaults']
        for i in defaults:
            DEFAULT_PARAMS_DICT.update({str(i): str(defaults[i])})


def is_any(value, needed_type):

    proof_list = []

    if needed_type == "float":

        try:

            proof_list.append(float(value))
            if float(value) < 0:

                return False

            else:

                return True

        except ValueError:

            return False

    elif needed_type == "float01":

        try:

            proof_list.append(float(value))
            if 1 > float(value) > 0:

                return True

            else:

                return False

        except ValueError:

            return False

    elif needed_type == "int":

        try:

            proof_list.append(int(value))
            if int(value) < 0:

                return False

            else:

                return True

        except ValueError:

            return False


# possible function to use selenium with proxy:
def get_proxy():

    pass

    # proxy_dict = {"20.210.113.32": "80", "71.86.129.131": "8080", "34.23.45.223": "80", "43.255.113.232": "80",
    #              "20.24.43.214": "80"}


# function to calculate dmarket/steam arbitrage if requested
# params (skin name, family_needed(if name is similar to others), skin_family(NULL by default)
def calculate_dmarket_steam_arbitrage(sn, fn, sf):

    # get dmarket cheapest price and wait
    dm_price = get_dmarket_html(sn, fn, sf)
    time.sleep(1)

    # get steam price
    st_price = get_steam_html(sn)
    time.sleep(1)

    # calculate arbitrage
    dmst_arbitrage = dm_price / float(st_price)

    # return value to spreadsheet
    return dmst_arbitrage


# function to get dmarket cheapest price:
# params (skin name, family_needed(if name is similar to others), skin_family(NULL by default)
def get_dmarket_html(skin_name, family_needed, skin_family):
    # check if you need the family of the skin
    if family_needed is True:

        # create dynamic url
        DMARKET_URL = f"https://dmarket.com/de/ingame-items/item-list/csgo-skins?family={skin_family}&title={skin_name}"

    else:

        # create dynamic url
        DMARKET_URL = f"https://dmarket.com/de/ingame-items/item-list/csgo-skins?title={skin_name}"

    # driver (currently Edge)
    driver = webdriver.Edge()

    # go to site and wait for page load
    driver.get(DMARKET_URL)
    time.sleep(5)

    # click buttons to filter by lowest price and wait for page load
    driver.find_element(By.CSS_SELECTOR, "div[class='o-select__sortTexts'] span").click()
    driver.find_element(By.CSS_SELECTOR, "button:nth-child(6) strong:nth-child(1)").click()
    time.sleep(WAIT_TIME)

    # get html_contents
    contents = BeautifulSoup(driver.page_source, "html.parser")

    # start function which will return the extracted price value
    return extract_dmarket_contents(contents)


# function to extract the price value out of the html contents
# params (html contents)
def extract_dmarket_contents(cts):
    # third translation tabel specific to this function to preserve some character
    translation_table = dict.fromkeys(map(ord, '$'), None)

    # extract all price, quality and float values from html contents
    # all_names = currently cant extract names
    all_prices = cts.find_all("price", class_="ng-star-inserted")
    all_qualities = cts.find_all("a", class_="c-asset__exterior c-asset__exterior--link ng-star-inserted")
    all_floats = cts.find_all("span", class_="o-blur")

    # create all intermittent lists
    # all_names_list = []

    all_prices_list = []
    all_qualities_list = []
    all_floats_list = []

    # strip and append values to intermittent lists
    # all prices
    for i in all_prices:
        all_prices_list.append(i.text.translate(translation_table))

    # all qualities
    for i in all_qualities:
        all_qualities_list.append(i.text)

    # all floats
    for i in all_floats:
        all_floats_list.append(i.text)

    # print all lists
    print(f"all prices(dmarket)"  # ignore the \xa (it's a comma)
          f"{all_prices_list}")

    print(f"all qualities"
          f"{all_qualities_list}")

    print(f"all floats"
          f"{all_floats_list}")

    # try to return a number, if it throws error do some formatting
    try:

        # return as float
        return float(all_prices_list[0])

    # catch ValueError
    except ValueError:

        # some fucked up shit that's not working
        new_int = str(all_prices_list[0][0]) + str(all_prices_list[0][2:-1])
        new_new_float = ""
        help_float_list = []

        for i in new_int:
            help_float_list.append(i)

        for k in help_float_list:
            new_new_float = new_new_float + k

        return float(new_new_float)


# function to get steam html
# params (skin name)
def get_steam_html(skin_name):
    # steam url
    STEAM_URL = f"https://steamcommunity.com/market/search?appid=730&q={skin_name}"

    # driver (currently Edge)
    driver = webdriver.Edge()

    # go to site and wait for page load
    driver.get(STEAM_URL)
    time.sleep(WAIT_TIME)

    # get html contents
    contents = BeautifulSoup(driver.page_source, "html.parser")

    # call function which will return the cheapest steam price
    return extract_steam_data(contents)


# function which extracts steam html contents
# params (html contents of page)
def extract_steam_data(cts):
    # third translation tabel specific to this function to preserve some character
    translation_table = dict.fromkeys(map(ord, '$USD\n\tarting: '), None)

    # extract all name and price values on page
    all_names = cts.find_all("span", class_="market_listing_item_name")
    all_prices = cts.find_all("span", class_="normal_price")

    # create all intermittent lists
    all_steam_names_list = []
    all_steam_prices_list = []

    # strip and append all items to intermittent lists
    for i in all_names:
        all_steam_names_list.append(i.text)

    for k in all_prices:
        all_steam_prices_list.append(k.text.translate(translation_table))

    # pop all duplicate price items
    for i in range(len(all_steam_prices_list) - 1, -1, -1):

        if abs(i) % 2 == 0:

            all_steam_prices_list.pop(i)
        else:
            pass

    # print all lists
    print(f"all names (steam)"
          f"{all_steam_names_list}")
    print(f"all prices (steam)"
          f"{all_steam_prices_list}")

    # return cheapest price
    return float(all_steam_prices_list[0])


# start main function for testing (CURRENT SOLUTION)



# open the log file now to ease future logging

    # main_func()


# calculate_dmarket_steam_arbitrage("AWP | Asiimov (Field-Tested)", False, "")

# if __name__ == "main":
#         main   ()
